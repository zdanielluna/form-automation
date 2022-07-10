import logging
import os
import sys
import shutil
import openpyxl
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC

logger = logging.getLogger('selenium.webdriver.remote.remote_connection')
logger.setLevel(logging.ERROR)
geckodriver_path = '[PATH]'
service = Service(geckodriver_path)
driver = webdriver.Firefox(service=service)

webpage = 'https://www.rpachallenge.com/'
wb_name = 'challenge.xlsx'
wb_path = os.path.join(os.getcwd(), wb_name)


def run(speed=0.20):
    try:
        workbook = openpyxl.load_workbook(wb_name)
    except:
        sys.exit(
            f'Ocorreu um erro ao acessar "{wb_name}". Verifique se o arquivo existe ou feche-o caso esteja aberto')

    sheet = workbook.active
    for row in range(2, sheet.max_row+1):
        data = []
        if sheet[f'A{row}'].value == None:
            break
        for column_index in range(1, sheet.max_column):
            column_letter = openpyxl.utils.cell.get_column_letter(column_index)
            cell = f'{column_letter}{row}'
            if sheet[cell].value != None:
                data.append(sheet[cell].value)

        insert_values(data, speed)


def insert_values(data, speed):
    try:
        role_in_company = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//label[text() = "Role in Company"]/following::input[1]')))
        email = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//label[text() = "Email"]/following::input[1]')))
        first_name = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//label[text() = "First Name"]/following::input[1]')))
        last_name = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//label[text() = "Last Name"]/following::input[1]')))
        company_name = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//label[text() = "Company Name"]/following::input[1]')))
        phone_number = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//label[text() = "Phone Number"]/following::input[1]')))
        address = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//label[text() = "Address"]/following::input[1]')))
        submit_button = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '/html/body/app-root/div[2]/app-rpa1/div/div[2]/form/input')))
    except:
        sys.exit('Os elementos da página não carregaram')

    first_name.send_keys(data[0])
    sleep(speed)
    last_name.send_keys(data[1])
    sleep(speed)
    company_name.send_keys(data[2])
    sleep(speed)
    role_in_company.send_keys(data[3])
    sleep(speed)
    address.send_keys(data[4])
    sleep(speed)
    email.send_keys(data[5])
    sleep(speed)
    phone_number.send_keys(data[6])
    sleep(speed)
    submit_button.click()


def download_wb():
    folder = 'Download'
    try:
        button = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((
                By.XPATH, f'/html/body/app-root/div[2]/app-rpa1/div/div[1]/div[6]/a')))
        button.click()
        return os.path.join(os.path.expanduser('~'), f'{folder}/{wb_name}')
    except:
        sys.exit(f'Ocorreu um erro ao tentar realizar o download')


def move_wb_to_project_folder(source):
    dest = os.getcwd()
    if os.path.exists(os.path.join(dest, wb_name)):
        os.remove(os.path.join(dest, wb_name))
    if os.path.exists(source):
        shutil.move(source, dest)


if __name__ == '__main__':
    try:
        driver.get(webpage)
        driver.maximize_window()
    except:
        sys.exit(f'Ocorreu um erro ao tentar acessar "{webpage}".')

    wb_download_path = download_wb()

    count = 60
    while not os.path.exists(wb_download_path):
        print(f'Aguardando arquivo, faltam {count} segundos...')
        count -= 1
        if count == 0:
            break
        sleep(1)
    else:
        print(f'Arquivo"{wb_download_path}" encontrado.')

    move_wb_to_project_folder(wb_download_path)
    run(speed=0.50)
    driver.quit()
